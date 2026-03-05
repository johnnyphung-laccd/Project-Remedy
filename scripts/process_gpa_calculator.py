#!/usr/bin/env python3
"""Process GPA Calculator XLSX through the LAMC ADA remediation pipeline."""

import sys
sys.path.insert(0, "src")

import asyncio
import hashlib
from pathlib import Path

import httpx

from lamc_adrp.config import load_config
from lamc_adrp.models import DocumentJob, FileType, JobStatus
from lamc_adrp.database import DatabaseManager
from lamc_adrp.zai_client import ZAIClient
from lamc_adrp.converter import HTMLConverter

# --- Constants ---
DOC_URL = "https://www.lamission.edu/sites/lamc.edu/files/2023-01/GPA%20Calculator2.xlsx"
SOURCE_PAGE = "https://lamission.edu/student-services/transfer-center/ASSIST"
LINK_TEXT = "GPA Calculator"
OUTPUT_HTML = Path("output/test_gpa_calculator.html")
DOWNLOAD_DIR = Path("output/downloads/xlsx")


async def main():
    # =====================================================================
    # Step 1: Download the XLSX
    # =====================================================================
    print("=" * 70)
    print("STEP 1: Downloading XLSX document")
    print("=" * 70)
    print(f"  URL: {DOC_URL}")

    async with httpx.AsyncClient(follow_redirects=True, verify=False, timeout=60.0) as client:
        response = await client.get(DOC_URL)
        response.raise_for_status()
        doc_bytes = response.content

    print(f"  Downloaded: {len(doc_bytes):,} bytes")
    print(f"  Status: {response.status_code}")
    print(f"  Content-Type: {response.headers.get('content-type', 'unknown')}")

    # =====================================================================
    # Step 2: Save with SHA-256 prefix
    # =====================================================================
    print("\n" + "=" * 70)
    print("STEP 2: Saving file with SHA-256 prefix")
    print("=" * 70)

    file_hash = hashlib.sha256(doc_bytes).hexdigest()
    filename = f"{file_hash[:16]}_GPA_Calculator2.xlsx"
    local_path = DOWNLOAD_DIR / filename
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    local_path.write_bytes(doc_bytes)

    print(f"  SHA-256: {file_hash}")
    print(f"  Saved to: {local_path}")
    print(f"  File size: {local_path.stat().st_size:,} bytes")

    # =====================================================================
    # Step 3: Extract content with openpyxl
    # =====================================================================
    print("\n" + "=" * 70)
    print("STEP 3: Native content extraction with openpyxl")
    print("=" * 70)

    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(str(local_path), data_only=False)
    print(f"  Sheets: {wb.sheetnames}")

    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  --- Processing sheet: '{sheet_name}' ---")
        print(f"      Dimensions: {ws.dimensions}")
        print(f"      Rows: {ws.max_row}, Columns: {ws.max_column}")

        # Add sheet name as heading
        parts.append(f"## {sheet_name}")

        # Collect merged cell info
        merged_ranges = list(ws.merged_cells.ranges)
        if merged_ranges:
            merge_notes = []
            for mr in merged_ranges:
                merge_notes.append(f"  - Merged: {mr}")
            parts.append("**Merged cells:**\n" + "\n".join(merge_notes))
            print(f"      Merged cells: {len(merged_ranges)}")

        # Read all rows into a list, tracking formulas and values
        all_rows = []
        formula_notes = []
        max_col = ws.max_column or 1
        max_row = ws.max_row or 1

        # Also load data_only version to get computed values
        wb_data = load_workbook(str(local_path), data_only=True)
        ws_data = wb_data[sheet_name]

        for row_idx in range(1, max_row + 1):
            row_cells = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_data = ws_data.cell(row=row_idx, column=col_idx)

                value = cell.value
                data_value = cell_data.value

                # Detect formulas
                if isinstance(value, str) and value.startswith("="):
                    col_letter = get_column_letter(col_idx)
                    cell_ref = f"{col_letter}{row_idx}"
                    # Use computed value if available, otherwise describe formula
                    if data_value is not None:
                        display = str(data_value)
                        formula_notes.append(
                            f"  - Cell {cell_ref}: formula `{value}` (computed: {data_value})"
                        )
                    else:
                        display = f"[Formula: {value}]"
                        formula_notes.append(
                            f"  - Cell {cell_ref}: formula `{value}`"
                        )
                elif value is not None:
                    display = str(value)
                else:
                    display = ""

                # Note formatting
                if cell.font and cell.font.bold:
                    if display and not display.startswith("**"):
                        display = f"**{display}**"

                row_cells.append(display.replace("|", "\\|").replace("\n", " "))

            all_rows.append(row_cells)

        wb_data.close()

        # Skip completely empty sheets
        has_content = any(any(c.strip() for c in row) for row in all_rows)
        if not has_content:
            parts.append("*(Empty sheet)*")
            print("      (empty sheet)")
            continue

        # Trim trailing empty rows
        while all_rows and not any(c.strip() for c in all_rows[-1]):
            all_rows.pop()

        # Trim trailing empty columns
        if all_rows:
            max_used_col = 0
            for row in all_rows:
                for i in range(len(row) - 1, -1, -1):
                    if row[i].strip():
                        max_used_col = max(max_used_col, i + 1)
                        break
            if max_used_col > 0:
                all_rows = [row[:max_used_col] for row in all_rows]

        # Build markdown table
        if all_rows:
            # First row as header
            header = all_rows[0]
            md_lines = ["| " + " | ".join(header) + " |"]
            md_lines.append("| " + " | ".join("---" for _ in header) + " |")
            for row in all_rows[1:]:
                # Pad row if shorter than header
                padded = row + [""] * (len(header) - len(row))
                md_lines.append("| " + " | ".join(padded[:len(header)]) + " |")
            parts.append("\n".join(md_lines))

            print(f"      Table: {len(all_rows)} rows x {len(header)} cols")

        # Add formula notes
        if formula_notes:
            parts.append("**Formulas in this sheet:**\n" + "\n".join(formula_notes))
            print(f"      Formulas detected: {len(formula_notes)}")

    wb.close()

    markdown_content = "\n\n".join(parts)

    print(f"\n  Total extracted markdown length: {len(markdown_content):,} chars")
    print(f"  Non-empty parts: {len(parts)}")

    print("\n  --- Extracted Markdown Preview (first 3000 chars) ---")
    print(markdown_content[:3000])
    print("  --- End Preview ---")

    # =====================================================================
    # Step 4-5: HTMLConverter - Plan and Generate accessible HTML
    # =====================================================================
    print("\n" + "=" * 70)
    print("STEP 4-5: Planning and generating accessible HTML via HTMLConverter")
    print("=" * 70)

    cfg = load_config()
    db = DatabaseManager(Path(cfg.output.db_path))
    await db.connect()

    zai = ZAIClient(cfg)
    await zai.start()

    try:
        # Create a DocumentJob
        job = DocumentJob(
            url=DOC_URL,
            source_page_url=SOURCE_PAGE,
            link_text=LINK_TEXT,
            link_context="Transfer Center - ASSIST page, GPA Calculator spreadsheet",
            file_type=FileType.XLSX,
            local_path=str(local_path),
            file_hash=file_hash,
            file_size=len(doc_bytes),
            status=JobStatus.EXTRACTED,
            ocr_markdown=markdown_content,
        )

        await db.create_job(job)
        print(f"  Job created: {job.id}")

        # Use HTMLConverter for planning + generation
        converter = HTMLConverter(cfg, zai, db)

        print("\n  Stage 4: Planning conversion...")
        plan = await converter.plan(job)
        print(f"  Plan generated: {len(plan):,} chars")
        print("\n  --- Conversion Plan Preview (first 1500 chars) ---")
        print(plan[:1500])
        print("  --- End Plan Preview ---")

        print("\n  Stage 5: Generating accessible HTML...")
        full_html = await converter.generate(job)
        print(f"  HTML generated: {len(full_html):,} chars")

        # =====================================================================
        # Step 6: Save output HTML
        # =====================================================================
        print("\n" + "=" * 70)
        print("STEP 6: Saving output HTML")
        print("=" * 70)

        OUTPUT_HTML.parent.mkdir(parents=True, exist_ok=True)
        OUTPUT_HTML.write_text(full_html, encoding="utf-8")
        print(f"  Saved to: {OUTPUT_HTML}")
        print(f"  File size: {OUTPUT_HTML.stat().st_size:,} bytes")

        # Print token usage
        print(f"\n  Token usage - input: {zai.total_input_tokens:,}, output: {zai.total_output_tokens:,}")
        print("\n  DONE!")

    finally:
        await zai.close()
        await db.close()


if __name__ == "__main__":
    asyncio.run(main())
